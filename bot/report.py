"""
Генерация Excel-отчёта с рекомендациями по ценам.

Логика:
1. Загрузка данных (30д, 14д, 7д) из WB API
2. Группировка товаров (A/B/C) по средним продажам
3. Расчёт avg_per_day по периоду группы
4. Расчёт days_remaining (на сколько дней хватит остатка)
5. Шкала повышения цены
6. Формирование Excel с 2 листами
"""

import os
import sys
import logging
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Настройка логирования
logger = logging.getLogger(__name__)

# Добавляем родительскую директорию для импорта wb_api
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from wb_api import get_orders, get_stocks, merge_orders_stocks, calc_avg_per_day
from bot.config import THRESHOLD_A, THRESHOLD_B, REPORTS_DIR


def assign_group(avg_per_day: float) -> str:
    """
    Определяет группу товара по средним продажам в день.

    A: ≥4 шт/день (ходовые) — анализ за 7 дней
    B: ≥0.5 шт/день — анализ за 14 дней
    C: <0.5 шт/день (редкие) — анализ за 30 дней
    """
    if avg_per_day >= THRESHOLD_A:
        return 'A'
    elif avg_per_day >= THRESHOLD_B:
        return 'B'
    else:
        return 'C'


def calc_avg_by_group(row) -> float:
    """
    Возвращает среднее продаж в день по периоду группы.

    A: по 7 дням
    B: по 14 дням
    C: по 30 дням
    """
    if row['group'] == 'A':
        return row['orders_count_7d'] / 7 if pd.notna(row['orders_count_7d']) else 0
    elif row['group'] == 'B':
        return row['orders_count_14d'] / 14 if pd.notna(row['orders_count_14d']) else 0
    else:  # C
        return row['avg_per_day_30d']


def calc_days_remaining(row) -> float:
    """
    На сколько дней хватит остатка.

    Returns:
        None если нет продаж, иначе кол-во дней
    """
    if row['avg_per_day'] == 0:
        return None
    return row['stock_qty'] / row['avg_per_day']


def get_price_increase(days_remaining: float) -> int:
    """
    Возвращает % повышения цены по шкале.

    Шкала:
    - > 7 дней: 0%
    - 6-7 дней: 5%
    - 5-6 дней: 10%
    - 4-5 дней: 15%
    - 3-4 дней: 25%
    - 2-3 дней: 30%
    - 1-2 дней: 40%
    - < 1 дня: 50%
    """
    if days_remaining is None or days_remaining > 7:
        return 0
    elif days_remaining >= 6:
        return 5
    elif days_remaining >= 5:
        return 10
    elif days_remaining >= 4:
        return 15
    elif days_remaining >= 3:
        return 25
    elif days_remaining >= 2:
        return 30
    elif days_remaining >= 1:
        return 40
    else:
        return 50


def format_worksheet(ws, column_widths: dict):
    """
    Форматирует лист Excel: ширина столбцов, шрифт, выравнивание.

    Args:
        ws: worksheet объект openpyxl
        column_widths: словарь {номер_столбца: ширина}
    """
    # Шрифт для заголовков (жирный)
    header_font = Font(name='Arial', size=11, bold=True)
    # Шрифт для данных
    data_font = Font(name='Arial', size=10)
    # Выравнивание по центру
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Форматируем заголовки (первая строка)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align

    # Форматируем данные
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            # Первый столбец (артикул) — по левому краю
            if cell.column == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Устанавливаем ширину столбцов
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


def generate_report() -> str:
    """
    Генерирует Excel-отчёт с рекомендациями по ценам.

    Returns:
        Путь к сгенерированному файлу
    """
    logger.info("=== Начало генерации отчёта ===")

    # === 1. Загрузка данных за 30 дней ===
    try:
        logger.info("Загрузка заказов за 30 дней...")
        orders_30d = get_orders(30)
        logger.info(f"✓ Заказы 30д: {len(orders_30d)} записей")
    except Exception as e:
        logger.error(f"✗ Ошибка загрузки заказов 30д: {e}")
        raise

    try:
        logger.info("Загрузка остатков...")
        stocks = get_stocks(orders_30d['nmId'].tolist())
        logger.info(f"✓ Остатки: {len(stocks)} записей")
    except Exception as e:
        logger.error(f"✗ Ошибка загрузки остатков: {e}")
        raise

    # Объединяем заказы и остатки
    df = merge_orders_stocks(orders_30d, stocks)
    df = calc_avg_per_day(df, days=30)

    # Заполняем поля возвратов если отсутствуют после merge
    for col in ['in_way_from_client', 'stock_qty_clean']:
        if col not in df.columns:
            df[col] = 0
        df[col] = df[col].fillna(0).astype(int)

    # Используем чистый остаток (без товаров в возврате) для расчётов
    df['stock_qty_original'] = df['stock_qty']
    df['stock_qty'] = df['stock_qty_clean']

    logger.info(f"Объединено товаров: {len(df)}")

    # === 2. Группировка товаров (A/B/C) ===
    df['group'] = df['avg_per_day_30d'].apply(assign_group)

    # === 3. Загрузка данных за 7 и 14 дней ===
    try:
        logger.info("Загрузка заказов за 7 дней...")
        orders_7d = get_orders(7)
        orders_7d = orders_7d[['nmId', 'orders_count_7d']]
        logger.info(f"✓ Заказы 7д: {len(orders_7d)} записей")
    except Exception as e:
        logger.error(f"✗ Ошибка загрузки заказов 7д: {e}")
        raise

    try:
        logger.info("Загрузка заказов за 14 дней...")
        orders_14d = get_orders(14)
        orders_14d = orders_14d[['nmId', 'orders_count_14d']]
        logger.info(f"✓ Заказы 14д: {len(orders_14d)} записей")
    except Exception as e:
        logger.error(f"✗ Ошибка загрузки заказов 14д: {e}")
        raise

    # Присоединяем к основной таблице
    df = df.merge(orders_7d, on='nmId', how='left')
    df = df.merge(orders_14d, on='nmId', how='left')

    # === 4. Расчёт среднего по группе ===
    df['avg_per_day'] = df.apply(calc_avg_by_group, axis=1)

    # === 5. Расчёт days_remaining ===
    df['days_remaining'] = df.apply(calc_days_remaining, axis=1)

    # === 6. Расчёт % повышения цены ===
    df['price_increase_pct'] = df['days_remaining'].apply(get_price_increase)

    # === 7. Формирование отчётов ===
    logger.info("Формирование отчётов...")

    # Лист 1: Товары для повышения цены (есть остаток И нужно повышение)
    report = df[(df['stock_qty'] > 0) & (df['price_increase_pct'] > 0)].copy()
    report = report.sort_values('days_remaining')
    logger.info(f"Товаров для повышения цены: {len(report)}")

    # Лист 2: Товары с нулевым остатком
    out_of_stock = df[df['stock_qty'] == 0].copy()
    out_of_stock = out_of_stock.sort_values(['group', 'avg_per_day'], ascending=[True, False])
    logger.info(f"Товаров с нулевым остатком: {len(out_of_stock)}")

    # === 8. Сохранение Excel ===
    os.makedirs(REPORTS_DIR, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_path = os.path.join(REPORTS_DIR, f'price_report_{timestamp}.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Лист 1: с колонкой "В возвратах" для прозрачности
        report_export = report[['supplierArticle', 'group', 'stock_qty',
                                'in_way_from_client',
                                'avg_per_day', 'days_remaining', 'price_increase_pct']]
        report_export.columns = ['Артикул', 'Группа', 'Остаток (чист.)',
                                 'В возвратах',
                                 'Продаж/день', 'Дней осталось', 'Повышение %']
        report_export.to_excel(writer, sheet_name='Повысить цену', index=False)

        # Лист 2
        out_of_stock_export = out_of_stock[['supplierArticle', 'group', 'avg_per_day']]
        out_of_stock_export.columns = ['Артикул', 'Группа', 'Продаж/день']
        out_of_stock_export.to_excel(writer, sheet_name='Нет на складе', index=False)

        # Добавляем итоги внизу листа 2
        ws2 = writer.sheets['Нет на складе']
        last_row = len(out_of_stock_export) + 3
        ws2.cell(row=last_row, column=1, value=f'Товаров с нулевым остатком: {len(out_of_stock)}')
        ws2.cell(row=last_row + 1, column=1, value=f'Упущенные продажи в день: {out_of_stock["avg_per_day"].sum():.1f} шт')

        # === 9. Форматирование Excel ===
        logger.info("Форматирование Excel...")

        # Лист 1: ширина столбцов
        ws1 = writer.sheets['Повысить цену']
        format_worksheet(ws1, {1: 20, 2: 10, 3: 16, 4: 14, 5: 14, 6: 16, 7: 14})

        # Лист 2: ширина столбцов
        format_worksheet(ws2, {1: 20, 2: 10, 3: 14})

        # === 10. Аннотации под таблицами ===
        # Лист 1: аннотация
        ws1_last = len(report_export) + 4
        ws1.cell(row=ws1_last, column=1, value='— Группы товаров —')
        ws1.cell(row=ws1_last + 1, column=1, value='A: ходовые (≥4 шт/день за 30д) → среднее по 7 дням')
        ws1.cell(row=ws1_last + 2, column=1, value='B: средние (≥0.5 шт/день) → среднее по 14 дням')
        ws1.cell(row=ws1_last + 3, column=1, value='C: редкие (<0.5 шт/день) → среднее по 30 дням')

        # Лист 2: аннотация (после итогов)
        ws2_last = last_row + 4
        ws2.cell(row=ws2_last, column=1, value='— Группы товаров —')
        ws2.cell(row=ws2_last + 1, column=1, value='A: ходовые (≥4 шт/день за 30д)')
        ws2.cell(row=ws2_last + 2, column=1, value='B: средние (≥0.5 шт/день)')
        ws2.cell(row=ws2_last + 3, column=1, value='C: редкие (<0.5 шт/день)')

    logger.info(f"✓ Отчёт сохранён: {output_path}")
    logger.info("=== Генерация завершена ===")

    return output_path
